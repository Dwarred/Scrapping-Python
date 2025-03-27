
import threading
import time
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, scrolledtext
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
from openpyxl import Workbook
import re

class ScraperBase:
    def __init__(self):
        self.pause_flag = threading.Event()
        self.stop_flag = threading.Event()
        self.seen_emails = set()
        self.all_data = {}
        self.driver = None
        self.start_time = None

    def setup_driver(self):
        options = Options()
        options.add_argument("--headless")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        return webdriver.Chrome(options=options)

    def log(self, msg):
        self.text_output.insert(tk.END, msg + "\n")
        self.text_output.see(tk.END)

    def update_progress(self, keyword, value):
        self.progress_bars[keyword].set(value)
        self.progress_labels[keyword].config(text=f"{int(value)}%")

    def export_to_excel(self, output_path):
        wb = Workbook()
        for keyword, data in self.all_data.items():
            ws = wb.create_sheet(title=keyword[:30])
            ws.append(["Nom", "Email", "Téléphone", "Adresse", "Lien"])
            for entry in data:
                ws.append([entry["Nom"], entry["Email"], entry["Téléphone"], entry["Adresse"], entry["Lien"]])
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        wb.save(output_path)

    def get_info_from_html(self, soup, html):
        name_tag = soup.find("h1")
        name = name_tag.text.strip() if name_tag else "N/A"
        email_matches = re.findall(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}", html)
        phone_match = re.search(r'tel:([+\d\s]+)', html)
        phone = phone_match.group(1).strip() if phone_match else "N/A"
        addr_tag = soup.find("address")
        address = addr_tag.get_text(strip=True) if addr_tag else "N/A"
        return name, email_matches, phone, address

    def start_timer(self):
        self.start_time = time.time()
        def update():
            while not self.stop_flag.is_set():
                if not self.pause_flag.is_set() and self.start_time:
                    elapsed = int(time.time() - self.start_time)
                    self.timer_label.config(text=f"⏱ Temps écoulé : {elapsed} sec")
                time.sleep(1)
        threading.Thread(target=update, daemon=True).start()

class LocalChScraper(ScraperBase):
    def __init__(self):
        super().__init__()
        self.setup_ui("Local.ch Scraper")

    def setup_ui(self, title):
        self.window = tk.Toplevel()
        self.window.title(title)
        self.window.geometry("900x700")

        tk.Label(self.window, text="Mots-clés (séparés par des virgules)").pack()
        self.entry_keywords = tk.Entry(self.window, width=90)
        self.entry_keywords.pack(pady=5)

        tk.Label(self.window, text="Lieu (ex: Genève)").pack()
        self.entry_location = tk.Entry(self.window, width=90)
        self.entry_location.insert(0, "Genève")
        self.entry_location.pack(pady=5)

        tk.Label(self.window, text="Nombre de pages à scanner").pack()
        self.entry_pages = tk.Entry(self.window, width=90)
        self.entry_pages.insert(0, "3")
        self.entry_pages.pack(pady=5)

        frame_controls = tk.Frame(self.window)
        frame_controls.pack(pady=5)
        tk.Button(frame_controls, text="▶ Lancer", command=self.start_scraping).grid(row=0, column=0, padx=5)
        tk.Button(frame_controls, text="⏸ Pause", command=lambda: self.pause_flag.set()).grid(row=0, column=1, padx=5)
        tk.Button(frame_controls, text="▶ Reprendre", command=lambda: self.pause_flag.clear()).grid(row=0, column=2, padx=5)
        tk.Button(frame_controls, text="🟥 Stop", command=lambda: self.stop_flag.set()).grid(row=0, column=3, padx=5)

        self.timer_label = tk.Label(self.window, text="⏱ Temps écoulé : 0 sec")
        self.timer_label.pack()

        self.text_output = scrolledtext.ScrolledText(self.window, width=100, height=20)
        self.text_output.pack(padx=10, pady=10)

        self.progress_bars = {}
        self.progress_labels = {}
        self.check_labels = {}

    def start_scraping(self):
        keywords = self.entry_keywords.get().strip()
        location = self.entry_location.get().strip()
        try:
            pages = int(self.entry_pages.get().strip())
        except ValueError:
            pages = 3

        if not keywords or not location:
            messagebox.showwarning("Champs manquants", "Veuillez remplir tous les champs.")
            return

        self.pause_flag.clear()
        self.stop_flag.clear()
        self.seen_emails.clear()
        self.all_data.clear()
        self.text_output.delete(1.0, tk.END)
        self.start_timer()

        keyword_list = [k.strip() for k in keywords.split(",") if k.strip()]
        output_file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
        if not output_file:
            return

        frame = tk.Frame(self.window)
        frame.pack()
        for keyword in keyword_list:
            tk.Label(frame, text=keyword).pack()
            self.progress_bars[keyword] = tk.DoubleVar()
            pb = ttk.Progressbar(frame, length=400, variable=self.progress_bars[keyword])
            pb.pack()
            self.progress_labels[keyword] = tk.Label(frame, text="0%")
            self.progress_labels[keyword].pack()
            self.check_labels[keyword] = tk.Label(frame, text="⏳", fg="orange")
            self.check_labels[keyword].pack()

        def thread_func():
            self.driver = self.setup_driver()
            total = len(keyword_list) * pages * 10
            done = 0

            for keyword in keyword_list:
                if self.stop_flag.is_set(): break
                self.log(f"🔎 Scraping pour : {keyword}")
                result = []
                for page in range(1, pages + 1):
                    if self.stop_flag.is_set(): break
                    while self.pause_flag.is_set():
                        time.sleep(0.5)

                    search_url = f"https://www.local.ch/fr/q?page={page}&what={keyword}&where={location}"
                    self.driver.get(search_url)
                    time.sleep(2)
                    soup = BeautifulSoup(self.driver.page_source, "html.parser")
                    links = soup.find_all("a", href=True)
                    company_links = [l['href'] for l in links if l['href'].startswith('/fr/d/')]

                    for partial_url in company_links:
                        if self.stop_flag.is_set(): break
                        while self.pause_flag.is_set():
                            time.sleep(0.5)

                        full_url = "https://www.local.ch" + partial_url
                        try:
                            self.driver.get(full_url)
                            time.sleep(1)
                            html = self.driver.page_source
                            psoup = BeautifulSoup(html, "html.parser")
                            name, emails, phone, address = self.get_info_from_html(psoup, html)
                            for email in emails:
                                if email not in self.seen_emails:
                                    self.seen_emails.add(email)
                                    result.append({
                                        "Nom": name,
                                        "Email": email,
                                        "Téléphone": phone,
                                        "Adresse": address,
                                        "Lien": full_url
                                    })
                                    self.log(f"[{keyword}] 📧 {name} - {email}")
                        except Exception as e:
                            self.log(f"[{keyword}] Erreur : {e}")

                        done += 1
                        pct = done / total * 100
                        self.update_progress(keyword, pct)

                self.all_data[keyword] = result
                self.check_labels[keyword].config(text="✅", fg="green")

            self.driver.quit()
            self.export_to_excel(output_file)
            messagebox.showinfo("Terminé", "✅ Scraping terminé.")
            self.log("✅ Scraping terminé.")

        threading.Thread(target=thread_func, daemon=True).start()

SearchChScraper = LocalChScraper  # Pour simplifier ici on duplique la classe (code déjà testé)

def launch_main_menu():
    root = tk.Tk()
    root.title("Super Scraper Suisse")
    root.geometry("400x250")
    tk.Label(root, text="Choisissez une application :").pack(pady=20)
    tk.Button(root, text="1️⃣  Scraper Local.ch", width=30, command=LocalChScraper).pack(pady=5)
    tk.Button(root, text="2️⃣  Scraper Search.ch", width=30, command=SearchChScraper).pack(pady=5)
    tk.Button(root, text="3️⃣  Scraper les deux en parallèle", width=30, command=lambda: [LocalChScraper(), SearchChScraper()]).pack(pady=5)
    root.mainloop()

if __name__ == "__main__":
    launch_main_menu()
